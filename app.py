from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
import os
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime, date
from zoneinfo import ZoneInfo
from openpyxl import Workbook, load_workbook
from io import BytesIO

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "pelog_secret_key")

DATABASE_URL = os.environ.get("DATABASE_URL")

if not DATABASE_URL:
    DATABASE_URL = "dbname=pelog user=postgres password=123456 host=localhost port=5432"


def conectar():
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


FUSO_BRASIL = ZoneInfo("America/Sao_Paulo")


def agora_brasil():
    return datetime.now(FUSO_BRASIL).replace(tzinfo=None)


USUARIOS = {
    "admin": {"senha": "admin123", "tipo": "admin"},
    "portaria": {"senha": "portaria123", "tipo": "portaria"},
    "pedro_cd": {"senha": "123", "tipo": "encarregado"},
    "jean_cd": {"senha": "123", "tipo": "encarregado"},
    "jalison_cross": {"senha": "123", "tipo": "encarregado"},
    "elaine_cross": {"senha": "123", "tipo": "encarregado"},
}


def setor_do_usuario(usuario):
    if usuario in ["jalison_cross", "elaine_cross"]:
        return "CROSS"

    if usuario in ["pedro_cd", "jean_cd"]:
        return "CD"

    return None


def is_admin():
    return session.get("tipo") == "admin"


def is_cd_autorizado():
    return setor_do_usuario(session.get("usuario")) == "CD"


def is_cross_autorizado():
    return setor_do_usuario(session.get("usuario")) == "CROSS"


def limpar_texto(valor):
    if valor is None:
        return None
    return str(valor).strip()


def converter_data(valor):
    if valor in [None, ""]:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    valor = str(valor).strip()

    formatos = [
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%y",
        "%Y/%m/%d"
    ]

    for formato in formatos:
        try:
            return datetime.strptime(valor, formato).date()
        except ValueError:
            pass

    return None


def converter_numero(valor):
    if valor in [None, ""]:
        return None

    if isinstance(valor, (int, float)):
        return valor

    valor = str(valor).strip()
    valor = valor.replace(".", "").replace(",", ".")

    try:
        return float(valor)
    except ValueError:
        return None


def normalizar_cabecalho(valor):
    if valor is None:
        return ""

    texto = str(valor).strip().lower()

    trocas = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c",
        "º": "",
        "°": "",
        "ª": "a",
        "³": "3",
        "(": " ",
        ")": " ",
        ".": "",
        "-": " ",
        "/": " ",
        "\\": " ",
        "_": " "
    }

    for antigo, novo in trocas.items():
        texto = texto.replace(antigo, novo)

    texto = " ".join(texto.split())
    return texto


MAPA_COLUNAS_EXCEL = {
    "centro": "centro",
    "descricao veiculo": "descricao_veiculo",
    "descrição veiculo": "descricao_veiculo",
    "nome cliente fornecedor": "nome_cliente_fornecedor",
    "pagador do frete": "nome_cliente_fornecedor",
    "pagador frete": "nome_cliente_fornecedor",
    "local": "local",
    "data remessa recebimento": "data_remessa_recebimento",
    "nr remessa recebimento": "nr_remessa_recebimento",
    "regiao": "regiao",
    "região": "regiao",
    "nome transportadora": "nome_transportadora",
    "data carregar": "data_carregar",
    "parceiro yd": "parceiro_yd",
    "nome transp subsequente": "nome_transp_subsequente",
    "denominacao parceiro yd": "denominacao_parceiro_yd",
    "denominação parceiro yd": "denominacao_parceiro_yd",
    "cod cliente recebedor fornecedor": "cod_cliente_recebedor_fornecedor",
    "condicao expedicao": "condicao_expedicao",
    "condição expedição": "condicao_expedicao",
    "data agenda entrega": "data_agenda_entrega",
    "material": "material",
    "peso liquido": "peso_liquido",
    "peso líquido": "peso_liquido",
    "qtde remessa": "qtde_remessa",
    "volume acumulado": "volume_acumulado",
    "n transporte": "numero_transporte",
    "nº transporte": "numero_transporte",
    "numero transporte": "numero_transporte",
    "data nfe": "data_nfe",
    "nr nfe": "numero_nfe",
    "nrº nfe": "numero_nfe",
    "numero nfe": "numero_nfe",
    "nr doc nfe": "numero_nfe",
    "numero dt subsequente": "numero_dt_subsequente",
    "n dt subsequente": "numero_dt_subsequente",
    "inf agenda entrega": "info_agenda_entrega",
    "info agenda entrega": "info_agenda_entrega",
    "data hora agendamento": "data_hora_agendamento",
    "data hora reagendamento": "data_hora_reagendamento",
    "numero pedido": "numero_pedido",
    "n pedido": "numero_pedido",
    "chave acesso": "chave_acesso",
    "chave de acesso de 44 posicoes": "chave_acesso",
    "documento vendas compras": "documento_vendas_compras",
    "data pedido": "data_pedido",
    "data transporte": "data_transporte",
    "caracteristica veiculo": "caracteristica_veiculo",
    "característica veiculo": "caracteristica_veiculo",
    "qtd teorica paletizacao convertida": "qtd_teorica_paletizacao_convertida",
    "qtd teórica paletização convertida": "qtd_teorica_paletizacao_convertida",
    "cliente pallet": "cliente_pallet",
    "placa composicao": "placa_composicao",
    "placa composição": "placa_composicao",
    "placa": "placa_simples_veiculo",
    "placa simples veiculo": "placa_simples_veiculo",
    "placa simples veículo": "placa_simples_veiculo",
    "nome motorista": "nome_motorista",
    "programacao 18 05 2026 motorista": "nome_motorista",
    "programação 18 05 2026 motorista": "nome_motorista",
    "h o r a r i o chegada": "data_hora_agendamento",

    "manifesto": "numero_transporte",
    "motorista": "nome_motorista",
    "cpf": "chave_acesso",
    "veiculo": "placa_simples_veiculo",
    "veículo": "placa_simples_veiculo",

    "chaves": "chaves",
    "altura": "altura",
    "largura": "largura",
}


COLUNAS_PROGRAMACAO = [
    "centro",
    "descricao_veiculo",
    "nome_cliente_fornecedor",
    "local",
    "data_remessa_recebimento",
    "nr_remessa_recebimento",
    "regiao",
    "nome_transportadora",
    "data_carregar",
    "parceiro_yd",
    "nome_transp_subsequente",
    "denominacao_parceiro_yd",
    "cod_cliente_recebedor_fornecedor",
    "condicao_expedicao",
    "data_agenda_entrega",
    "material",
    "peso_liquido",
    "qtde_remessa",
    "volume_acumulado",
    "numero_transporte",
    "data_nfe",
    "numero_nfe",
    "numero_dt_subsequente",
    "info_agenda_entrega",
    "data_hora_agendamento",
    "data_hora_reagendamento",
    "numero_pedido",
    "chave_acesso",
    "documento_vendas_compras",
    "data_pedido",
    "data_transporte",
    "caracteristica_veiculo",
    "qtd_teorica_paletizacao_convertida",
    "cliente_pallet",
    "placa_composicao",
    "placa_simples_veiculo",
    "nome_motorista",
    "chaves",
    "altura",
    "largura",
]


COLUNAS_DATA = [
    "data_remessa_recebimento",
    "data_carregar",
    "data_agenda_entrega",
    "data_nfe",
    "data_pedido",
    "data_transporte",
]


COLUNAS_NUMERO = [
    "peso_liquido",
    "qtde_remessa",
    "volume_acumulado",
    "qtd_teorica_paletizacao_convertida",
]


COLUNAS_PROGRAMACAO_CD = [
    "numero_transporte",
    "nome_transportadora",
    "descricao_veiculo",
    "numero_nfe",
    "data_nfe",
    "chave_acesso",
    "data_agenda_entrega",
]


MAPA_COLUNAS_EXPEDICAO = {
    "destino": "destino",
    "cliente": "cliente",
    "material": "material",
    "nf": "nf",
    "dt 1 perna": "dt_primeira_perna",
    "dt 1a perna": "dt_primeira_perna",
    "dt 1Âª perna": "dt_primeira_perna",
    "dt 2 perna": "dt_segunda_perna",
    "dt 2a perna": "dt_segunda_perna",
    "dt 2Âª perna": "dt_segunda_perna",
    "inf agenda entrega": "info_agenda_entrega",
    "info agenda entrega": "info_agenda_entrega",
    "plt": "plt",
    "volume mÂ³": "volume_m3",
    "volume m3": "volume_m3",
    "peso ton": "peso_ton",
    "tipo de veiculo": "tipo_veiculo",
    "tipo de veÃ­culo": "tipo_veiculo",
    "altura plt": "altura_plt",
    "largura plt": "largura_plt",
    "veiculo": "veiculo",
    "veÃ­culo": "veiculo",
    "motorista": "motorista",
}


COLUNAS_EXPEDICAO = [
    "destino",
    "cliente",
    "material",
    "nf",
    "dt_primeira_perna",
    "dt_segunda_perna",
    "info_agenda_entrega",
    "plt",
    "volume_m3",
    "peso_ton",
    "tipo_veiculo",
    "altura_plt",
    "largura_plt",
    "veiculo",
    "motorista",
]


COLUNAS_NUMERO_EXPEDICAO = [
    "plt",
    "volume_m3",
    "peso_ton",
    "altura_plt",
    "largura_plt",
]


def criar_tabela_programacao(cur, nome_tabela):
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {nome_tabela} (
            id SERIAL PRIMARY KEY,
            centro VARCHAR(50),
            descricao_veiculo VARCHAR(150),
            nome_cliente_fornecedor VARCHAR(200),
            local VARCHAR(150),
            data_remessa_recebimento DATE,
            nr_remessa_recebimento VARCHAR(80),
            regiao VARCHAR(50),
            nome_transportadora VARCHAR(200),
            data_carregar DATE,
            parceiro_yd VARCHAR(100),
            nome_transp_subsequente VARCHAR(200),
            denominacao_parceiro_yd VARCHAR(200),
            cod_cliente_recebedor_fornecedor VARCHAR(80),
            condicao_expedicao VARCHAR(80),
            data_agenda_entrega DATE,
            material VARCHAR(100),
            peso_liquido NUMERIC(12,3),
            qtde_remessa NUMERIC(12,3),
            volume_acumulado NUMERIC(12,3),
            numero_transporte VARCHAR(80),
            data_nfe DATE,
            numero_nfe VARCHAR(80),
            numero_dt_subsequente VARCHAR(80),
            info_agenda_entrega TEXT,
            data_hora_agendamento VARCHAR(100),
            data_hora_reagendamento VARCHAR(100),
            numero_pedido VARCHAR(80),
            chave_acesso VARCHAR(80),
            documento_vendas_compras VARCHAR(80),
            data_pedido DATE,
            data_transporte DATE,
            caracteristica_veiculo VARCHAR(100),
            qtd_teorica_paletizacao_convertida NUMERIC(12,3),
            cliente_pallet VARCHAR(100),
            placa_composicao VARCHAR(80),
            placa_simples_veiculo VARCHAR(80),
            nome_motorista VARCHAR(200),
            chaves VARCHAR(150),
            altura VARCHAR(50),
            largura VARCHAR(50),
            horario_chegada TIMESTAMP,
            usuario_chegada VARCHAR(100),
            horario_saida TIMESTAMP,
            usuario_saida VARCHAR(100),
            justificativa TEXT,
            usuario_justificativa VARCHAR(100),
            horario_justificativa TIMESTAMP,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)

    cur.execute(f"ALTER TABLE {nome_tabela} ADD COLUMN IF NOT EXISTS horario_chegada TIMESTAMP;")
    cur.execute(f"ALTER TABLE {nome_tabela} ADD COLUMN IF NOT EXISTS usuario_chegada VARCHAR(100);")
    cur.execute(f"ALTER TABLE {nome_tabela} ADD COLUMN IF NOT EXISTS horario_saida TIMESTAMP;")
    cur.execute(f"ALTER TABLE {nome_tabela} ADD COLUMN IF NOT EXISTS usuario_saida VARCHAR(100);")
    cur.execute(f"ALTER TABLE {nome_tabela} ADD COLUMN IF NOT EXISTS justificativa TEXT;")
    cur.execute(f"ALTER TABLE {nome_tabela} ADD COLUMN IF NOT EXISTS usuario_justificativa VARCHAR(100);")
    cur.execute(f"ALTER TABLE {nome_tabela} ADD COLUMN IF NOT EXISTS horario_justificativa TIMESTAMP;")


def criar_tabela_expedicao(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS programacao_expedicao (
            id SERIAL PRIMARY KEY,
            destino VARCHAR(150),
            cliente VARCHAR(200),
            material VARCHAR(100),
            nf VARCHAR(100),
            dt_primeira_perna VARCHAR(100),
            dt_segunda_perna VARCHAR(100),
            info_agenda_entrega TEXT,
            plt NUMERIC(12,3),
            volume_m3 NUMERIC(12,3),
            peso_ton NUMERIC(12,3),
            tipo_veiculo VARCHAR(150),
            altura_plt NUMERIC(12,3),
            largura_plt NUMERIC(12,3),
            veiculo VARCHAR(100),
            motorista VARCHAR(200),
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)

    tipos = {
        "destino": "VARCHAR(150)",
        "cliente": "VARCHAR(200)",
        "material": "VARCHAR(100)",
        "nf": "VARCHAR(100)",
        "dt_primeira_perna": "VARCHAR(100)",
        "dt_segunda_perna": "VARCHAR(100)",
        "info_agenda_entrega": "TEXT",
        "plt": "NUMERIC(12,3)",
        "volume_m3": "NUMERIC(12,3)",
        "peso_ton": "NUMERIC(12,3)",
        "tipo_veiculo": "VARCHAR(150)",
        "altura_plt": "NUMERIC(12,3)",
        "largura_plt": "NUMERIC(12,3)",
        "veiculo": "VARCHAR(100)",
        "motorista": "VARCHAR(200)",
    }

    for coluna, tipo in tipos.items():
        cur.execute(f"ALTER TABLE programacao_expedicao ADD COLUMN IF NOT EXISTS {coluna} {tipo};")


def criar_colunas():
    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS caminhoes (
            id SERIAL PRIMARY KEY
        );
    """)

    colunas = [
        "ADD COLUMN IF NOT EXISTS placa VARCHAR(20)",
        "ADD COLUMN IF NOT EXISTS motorista VARCHAR(100)",
        "ADD COLUMN IF NOT EXISTS cpf VARCHAR(20)",
        "ADD COLUMN IF NOT EXISTS empresa VARCHAR(100)",
        "ADD COLUMN IF NOT EXISTS tipo_material VARCHAR(100)",
        "ADD COLUMN IF NOT EXISTS nota_fiscal VARCHAR(100)",
        "ADD COLUMN IF NOT EXISTS setor_doca VARCHAR(20)",
        "ADD COLUMN IF NOT EXISTS horario TIMESTAMP",
        "ADD COLUMN IF NOT EXISTS doca VARCHAR(50)",
        "ADD COLUMN IF NOT EXISTS status VARCHAR(50)",
        "ADD COLUMN IF NOT EXISTS autorizado_por VARCHAR(100)",
        "ADD COLUMN IF NOT EXISTS horario_autorizacao TIMESTAMP",
        "ADD COLUMN IF NOT EXISTS inicio_doca TIMESTAMP",
        "ADD COLUMN IF NOT EXISTS fim_doca TIMESTAMP",
        "ADD COLUMN IF NOT EXISTS dts_observacao TEXT",
        "ADD COLUMN IF NOT EXISTS produtos_sku VARCHAR(255)",
        "ADD COLUMN IF NOT EXISTS notas VARCHAR(255)",
        "ADD COLUMN IF NOT EXISTS diferenca_os VARCHAR(255)",
        "ADD COLUMN IF NOT EXISTS quantidade_nfs INTEGER",
        "ADD COLUMN IF NOT EXISTS qtd_paletes_nf INTEGER",
        "ADD COLUMN IF NOT EXISTS qtd_paletes_conferido INTEGER",
        "ADD COLUMN IF NOT EXISTS peso_kg NUMERIC(10,2)",
        "ADD COLUMN IF NOT EXISTS diferenca_produtos TEXT",
        "ADD COLUMN IF NOT EXISTS equipe TEXT",
        "ADD COLUMN IF NOT EXISTS operacional_cadastrado_por VARCHAR(100)",
        "ADD COLUMN IF NOT EXISTS horario_operacional TIMESTAMP",
        "ADD COLUMN IF NOT EXISTS horario_saida TIMESTAMP",
        "ADD COLUMN IF NOT EXISTS saida_registrada_por VARCHAR(100)"
    ]

    for coluna in colunas:
        cur.execute(f"ALTER TABLE caminhoes {coluna};")

    criar_tabela_programacao(cur, "programacao_cd")
    criar_tabela_programacao(cur, "programacao_cross")
    criar_tabela_expedicao(cur)

    conn.commit()
    cur.close()
    conn.close()


TABELAS_CRIADAS = False


@app.before_request
def iniciar():
    global TABELAS_CRIADAS

    if not TABELAS_CRIADAS:
        criar_colunas()
        TABELAS_CRIADAS = True


def importar_planilha_programacao(arquivo, tabela):
    conn = None
    cur = None

    try:
        wb = load_workbook(
            arquivo,
            data_only=True,
            read_only=True
        )

        ws = wb.active

        cabecalhos = []

        for celula in ws[1]:
            cabecalho_normalizado = normalizar_cabecalho(celula.value)
            coluna_banco = MAPA_COLUNAS_EXCEL.get(cabecalho_normalizado)

            if not coluna_banco and cabecalho_normalizado.endswith("motorista"):
                coluna_banco = "nome_motorista"

            cabecalhos.append(coluna_banco)

        registros_importados = 0
        contador_commit = 0

        conn = conectar()
        cur = conn.cursor()

        for linha in ws.iter_rows(min_row=2, values_only=True):
            dados = {}

            for indice, valor in enumerate(linha):
                if indice < len(cabecalhos):
                    coluna_banco = cabecalhos[indice]

                    if coluna_banco:
                        if coluna_banco in COLUNAS_DATA:
                            dados[coluna_banco] = converter_data(valor)
                        elif coluna_banco in COLUNAS_NUMERO:
                            dados[coluna_banco] = converter_numero(valor)
                        else:
                            dados[coluna_banco] = limpar_texto(valor)

            if not any(dados.values()):
                continue

            colunas = list(dados.keys())

            if not colunas:
                continue

            valores = [dados[coluna] for coluna in colunas]
            placeholders = ", ".join(["%s"] * len(colunas))

            sql = f"""
                INSERT INTO {tabela}
                ({", ".join(colunas)})
                VALUES ({placeholders});
            """

            cur.execute(sql, valores)

            registros_importados += 1
            contador_commit += 1

            if contador_commit >= 100:
                conn.commit()
                contador_commit = 0

        conn.commit()

        try:
            wb.close()
        except Exception:
            pass

        return registros_importados

    except Exception as e:
        if conn:
            conn.rollback()
        raise e

    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


def importar_planilha_expedicao(arquivo):
    conn = None
    cur = None

    try:
        wb = load_workbook(
            arquivo,
            data_only=True,
            read_only=True
        )

        ws = wb.active
        cabecalhos = []

        for celula in ws[1]:
            cabecalho_normalizado = normalizar_cabecalho(celula.value)
            cabecalhos.append(MAPA_COLUNAS_EXPEDICAO.get(cabecalho_normalizado))

        registros_importados = 0
        contador_commit = 0

        conn = conectar()
        cur = conn.cursor()

        for linha in ws.iter_rows(min_row=2, values_only=True):
            dados = {}

            for indice, valor in enumerate(linha):
                if indice < len(cabecalhos):
                    coluna_banco = cabecalhos[indice]

                    if coluna_banco:
                        if coluna_banco in COLUNAS_NUMERO_EXPEDICAO:
                            dados[coluna_banco] = converter_numero(valor)
                        else:
                            dados[coluna_banco] = limpar_texto(valor)

            if not any(dados.values()):
                continue

            colunas = list(dados.keys())

            if not colunas:
                continue

            valores = [dados[coluna] for coluna in colunas]
            placeholders = ", ".join(["%s"] * len(colunas))

            sql = f"""
                INSERT INTO programacao_expedicao
                ({", ".join(colunas)})
                VALUES ({placeholders});
            """

            cur.execute(sql, valores)

            registros_importados += 1
            contador_commit += 1

            if contador_commit >= 100:
                conn.commit()
                contador_commit = 0

        conn.commit()

        try:
            wb.close()
        except Exception:
            pass

        return registros_importados

    except Exception as e:
        if conn:
            conn.rollback()
        raise e

    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


def buscar_programacoes(tabela, busca="", data_inicio="", data_fim=""):
    conn = conectar()
    cur = conn.cursor()

    query = f"""
        SELECT *
        FROM {tabela}
        WHERE 1=1
    """

    params = []

    if busca:
        query += """
            AND (
                nome_cliente_fornecedor ILIKE %s OR
                local ILIKE %s OR
                nome_transportadora ILIKE %s OR
                descricao_veiculo ILIKE %s OR
                material ILIKE %s OR
                numero_nfe ILIKE %s OR
                numero_transporte ILIKE %s OR
                chave_acesso ILIKE %s OR
                placa_composicao ILIKE %s OR
                placa_simples_veiculo ILIKE %s OR
                nome_motorista ILIKE %s
            )
        """
        termo = f"%{busca}%"
        params.extend([termo] * 11)

    if data_inicio:
        query += " AND data_agenda_entrega >= %s"
        params.append(data_inicio)

    if data_fim:
        query += " AND data_agenda_entrega <= %s"
        params.append(data_fim)

    query += " ORDER BY data_agenda_entrega ASC NULLS LAST, id DESC;"

    cur.execute(query, params)
    dados = cur.fetchall()

    cur.close()
    conn.close()

    return dados


def buscar_programacao_expedicao(busca="", data_inicio="", data_fim=""):
    conn = conectar()
    cur = conn.cursor()

    query = """
        SELECT *
        FROM programacao_expedicao
        WHERE 1=1
    """

    params = []

    if busca:
        query += """
            AND (
                destino ILIKE %s OR
                cliente ILIKE %s OR
                material ILIKE %s OR
                nf ILIKE %s OR
                dt_primeira_perna ILIKE %s OR
                dt_segunda_perna ILIKE %s OR
                info_agenda_entrega ILIKE %s OR
                tipo_veiculo ILIKE %s OR
                veiculo ILIKE %s OR
                motorista ILIKE %s
            )
        """
        termo = f"%{busca}%"
        params.extend([termo] * 10)

    if data_inicio:
        query += " AND criado_em::date >= %s"
        params.append(data_inicio)

    if data_fim:
        query += " AND criado_em::date <= %s"
        params.append(data_fim)

    query += " ORDER BY id DESC;"

    cur.execute(query, params)
    dados = cur.fetchall()

    cur.close()
    conn.close()

    return dados


@app.route("/")
def index():
    if "usuario" not in session:
        return redirect(url_for("login"))

    if session.get("tipo") == "admin":
        return redirect(url_for("admin"))

    if session.get("tipo") == "portaria":
        return redirect(url_for("portaria"))

    if session.get("tipo") == "encarregado":
        return redirect(url_for("encarregado"))

    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form.get("usuario")
        senha = request.form.get("senha")

        if usuario in USUARIOS and USUARIOS[usuario]["senha"] == senha:
            session["usuario"] = usuario
            session["tipo"] = USUARIOS[usuario]["tipo"]
            return redirect(url_for("index"))

        return render_template("login.html", erro="Usuário ou senha inválidos")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/portaria")
def portaria():
    if session.get("tipo") != "portaria":
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM caminhoes
        WHERE status IS NULL OR status != 'finalizado'
        ORDER BY id DESC;
    """)

    caminhoes = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "portaria.html",
        caminhoes=caminhoes,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/api/portaria")
def api_portaria():
    if session.get("tipo") != "portaria":
        return jsonify([])

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            id, placa, motorista, empresa, tipo_material, nota_fiscal,
            setor_doca, status, doca
        FROM caminhoes
        WHERE status IS NULL OR status != 'finalizado'
        ORDER BY id DESC;
    """)

    caminhoes = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify(caminhoes)


@app.route("/registrar", methods=["POST"])
def registrar():
    if session.get("tipo") != "portaria":
        return redirect(url_for("login"))

    setor_doca = request.form.get("setor_doca")

    if setor_doca not in ["CROSS", "CD"]:
        return redirect(url_for("portaria"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO caminhoes
        (
            placa, motorista, cpf, empresa, tipo_material,
            nota_fiscal, setor_doca, horario, status
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'aguardando');
    """, (
        request.form.get("placa"),
        request.form.get("motorista"),
        request.form.get("cpf"),
        request.form.get("empresa"),
        request.form.get("tipo_material"),
        request.form.get("nota_fiscal"),
        setor_doca,
        agora_brasil()
    ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("portaria"))


@app.route("/encarregado")
def encarregado():
    if session.get("tipo") != "encarregado":
        return redirect(url_for("login"))

    usuario = session.get("usuario")
    setor = setor_do_usuario(usuario)

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM caminhoes
        WHERE setor_doca = %s
          AND (status IS NULL OR status != 'finalizado')
        ORDER BY id DESC;
    """, (setor,))

    caminhoes = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "encarregado.html",
        caminhoes=caminhoes,
        usuario=usuario,
        perfil=session.get("tipo"),
        setor=setor,
        pode_ver_cd=is_cd_autorizado(),
        pode_ver_cross=is_cross_autorizado()
    )


@app.route("/autorizar/<int:id>", methods=["POST"])
def autorizar(id):
    if session.get("tipo") != "encarregado":
        return redirect(url_for("login"))

    usuario = session.get("usuario")
    setor = setor_do_usuario(usuario)

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        UPDATE caminhoes
        SET doca = %s,
            status = 'autorizado',
            autorizado_por = %s,
            horario_autorizacao = %s
        WHERE id = %s
          AND setor_doca = %s;
    """, (
        request.form.get("doca"),
        usuario,
        agora_brasil(),
        id,
        setor
    ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("encarregado"))


@app.route("/iniciar_doca/<int:id>", methods=["POST"])
def iniciar_doca(id):
    if session.get("tipo") != "encarregado":
        return redirect(url_for("login"))

    setor = setor_do_usuario(session.get("usuario"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        UPDATE caminhoes
        SET status = 'na_doca',
            inicio_doca = %s
        WHERE id = %s
          AND setor_doca = %s;
    """, (
        agora_brasil(),
        id,
        setor
    ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("encarregado"))


@app.route("/finalizar_doca/<int:id>", methods=["POST"])
def finalizar_doca(id):
    if session.get("tipo") != "encarregado":
        return redirect(url_for("login"))

    setor = setor_do_usuario(session.get("usuario"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        UPDATE caminhoes
        SET status = 'doca_finalizada',
            fim_doca = %s
        WHERE id = %s
          AND setor_doca = %s;
    """, (
        agora_brasil(),
        id,
        setor
    ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("encarregado"))


@app.route("/finalizar_saida/<int:id>", methods=["POST"])
def finalizar_saida(id):
    if session.get("tipo") != "portaria":
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        UPDATE caminhoes
        SET status = 'finalizado',
            horario_saida = %s,
            saida_registrada_por = %s
        WHERE id = %s
          AND status = 'doca_finalizada';
    """, (
        agora_brasil(),
        session.get("usuario"),
        id
    ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("portaria"))


@app.route("/cadastro_operacional/<int:id>", methods=["GET", "POST"])
def cadastro_operacional(id):
    if session.get("tipo") != "encarregado":
        return redirect(url_for("login"))

    usuario = session.get("usuario")
    setor = setor_do_usuario(usuario)

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        cur.execute("""
            UPDATE caminhoes
            SET dts_observacao = %s,
                produtos_sku = %s,
                notas = %s,
                quantidade_nfs = %s,
                qtd_paletes_nf = %s,
                qtd_paletes_conferido = %s,
                peso_kg = %s,
                diferenca_produtos = %s,
                equipe = %s,
                operacional_cadastrado_por = %s,
                horario_operacional = %s
            WHERE id = %s
              AND setor_doca = %s;
        """, (
            request.form.get("dts_observacao"),
            request.form.get("produtos_sku"),
            request.form.get("notas"),
            request.form.get("quantidade_nfs") or None,
            request.form.get("qtd_paletes_nf") or None,
            request.form.get("qtd_paletes_conferido") or None,
            request.form.get("peso_kg") or None,
            request.form.get("diferenca_produtos"),
            request.form.get("equipe"),
            usuario,
            agora_brasil(),
            id,
            setor
        ))

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for("encarregado"))

    cur.execute("""
        SELECT * 
        FROM caminhoes 
        WHERE id = %s
          AND setor_doca = %s;
    """, (id, setor))

    caminhao = cur.fetchone()

    cur.close()
    conn.close()

    if not caminhao:
        return redirect(url_for("encarregado"))

    return render_template(
        "cadastro_operacional.html",
        caminhao=caminhao,
        registro=caminhao,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/admin")
def admin():
    if session.get("tipo") != "admin":
        return redirect(url_for("login"))

    busca = request.args.get("busca", "")
    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")

    conn = conectar()
    cur = conn.cursor()

    query = """
        SELECT *
        FROM caminhoes
        WHERE 1=1
    """

    params = []

    if busca:
        query += """
            AND (
                placa ILIKE %s OR
                motorista ILIKE %s OR
                cpf ILIKE %s OR
                empresa ILIKE %s OR
                tipo_material ILIKE %s OR
                nota_fiscal ILIKE %s OR
                setor_doca ILIKE %s OR
                doca ILIKE %s OR
                equipe ILIKE %s
            )
        """
        termo = f"%{busca}%"
        params.extend([termo, termo, termo, termo, termo, termo, termo, termo, termo])

    if data_inicio:
        query += " AND horario::date >= %s"
        params.append(data_inicio)

    if data_fim:
        query += " AND horario::date <= %s"
        params.append(data_fim)

    query += " ORDER BY id DESC;"

    cur.execute(query, params)
    caminhoes = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "admin.html",
        caminhoes=caminhoes,
        busca=busca,
        data_inicio=data_inicio,
        data_fim=data_fim,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/admin/programacao-cd", methods=["GET", "POST"])
def admin_programacao_cd():
    if not is_admin():
        return redirect(url_for("login"))

    mensagem = None
    erro = None

    if request.method == "POST":
        arquivo = request.files.get("arquivo_excel")

        if not arquivo or arquivo.filename == "":
            erro = "Selecione uma planilha Excel para importar."
        else:
            try:
                registros = importar_planilha_programacao(arquivo, "programacao_cd")
                mensagem = f"Planilha CD importada com sucesso. {registros} registros adicionados."
            except Exception as e:
                erro = f"Erro ao importar planilha CD: {str(e)}"

    busca = request.args.get("busca", "")
    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")

    programacoes = buscar_programacoes("programacao_cd", busca, data_inicio, data_fim)

    return render_template(
        "admin_programacao_cd.html",
        programacoes=programacoes,
        busca=busca,
        data_inicio=data_inicio,
        data_fim=data_fim,
        mensagem=mensagem,
        erro=erro,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/admin/programacao-cross", methods=["GET", "POST"])
def admin_programacao_cross():
    if not is_admin():
        return redirect(url_for("login"))

    mensagem = None
    erro = None

    if request.method == "POST":
        arquivo = request.files.get("arquivo_excel")

        if not arquivo or arquivo.filename == "":
            erro = "Selecione uma planilha Excel para importar."
        else:
            try:
                registros = importar_planilha_programacao(arquivo, "programacao_cross")
                mensagem = f"Planilha CROSS importada com sucesso. {registros} registros adicionados."
            except Exception as e:
                erro = f"Erro ao importar planilha CROSS: {str(e)}"

    busca = request.args.get("busca", "")
    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")

    programacoes = buscar_programacoes("programacao_cross", busca, data_inicio, data_fim)

    return render_template(
        "admin_programacao_cross.html",
        programacoes=programacoes,
        busca=busca,
        data_inicio=data_inicio,
        data_fim=data_fim,
        mensagem=mensagem,
        erro=erro,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/admin/programacao-expedicao", methods=["GET", "POST"])
def admin_programacao_expedicao():
    if not is_admin():
        return redirect(url_for("login"))

    mensagem = None
    erro = None

    if request.method == "POST":
        arquivo = request.files.get("arquivo_excel")

        if not arquivo or arquivo.filename == "":
            erro = "Selecione uma planilha Excel para importar."
        else:
            try:
                registros = importar_planilha_expedicao(arquivo)
                mensagem = f"Planilha Expedição importada com sucesso. {registros} registros adicionados."
            except Exception as e:
                erro = f"Erro ao importar planilha Expedição: {str(e)}"

    busca = request.args.get("busca", "")
    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")

    programacoes = buscar_programacao_expedicao(busca, data_inicio, data_fim)

    return render_template(
        "admin_programacao_expedicao.html",
        programacoes=programacoes,
        busca=busca,
        data_inicio=data_inicio,
        data_fim=data_fim,
        mensagem=mensagem,
        erro=erro,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/admin/programacao-cd/editar/<int:id>", methods=["GET", "POST"])
def editar_programacao_cd(id):
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        valores = []

        for coluna in COLUNAS_PROGRAMACAO_CD:
            valor = request.form.get(coluna)

            if coluna in COLUNAS_DATA:
                valor = converter_data(valor)
            elif coluna in COLUNAS_NUMERO:
                valor = converter_numero(valor)
            else:
                valor = limpar_texto(valor)

            valores.append(valor)

        valores.append(id)

        set_sql = ", ".join([f"{coluna} = %s" for coluna in COLUNAS_PROGRAMACAO_CD])

        cur.execute(f"""
            UPDATE programacao_cd
            SET {set_sql}
            WHERE id = %s;
        """, valores)

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for("admin_programacao_cd"))

    cur.execute("""
        SELECT *
        FROM programacao_cd
        WHERE id = %s;
    """, (id,))

    registro = cur.fetchone()

    cur.close()
    conn.close()

    if not registro:
        return redirect(url_for("admin_programacao_cd"))

    return render_template(
        "editar_programacao_cd.html",
        registro=registro,
        colunas=COLUNAS_PROGRAMACAO_CD,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/admin/programacao-cross/editar/<int:id>", methods=["GET", "POST"])
def editar_programacao_cross(id):
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        colunas_cross = [
            "numero_transporte",
            "nome_motorista",
            "nome_cliente_fornecedor",
            "placa_simples_veiculo",
        ]
        valores = []

        for coluna in colunas_cross:
            valores.append(limpar_texto(request.form.get(coluna)))

        valores.append(id)

        set_sql = ", ".join([f"{coluna} = %s" for coluna in colunas_cross])

        cur.execute(f"""
            UPDATE programacao_cross
            SET {set_sql}
            WHERE id = %s;
        """, valores)

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for("admin_programacao_cross"))

    cur.execute("""
        SELECT *
        FROM programacao_cross
        WHERE id = %s;
    """, (id,))

    registro = cur.fetchone()

    cur.close()
    conn.close()

    if not registro:
        return redirect(url_for("admin_programacao_cross"))

    return render_template(
        "editar_programacao_cross.html",
        registro=registro,
        colunas=COLUNAS_PROGRAMACAO,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/admin/programacao-expedicao/editar/<int:id>", methods=["GET", "POST"])
def editar_programacao_expedicao(id):
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        valores = []

        for coluna in COLUNAS_EXPEDICAO:
            valor = request.form.get(coluna)

            if coluna in COLUNAS_NUMERO_EXPEDICAO:
                valor = converter_numero(valor)
            else:
                valor = limpar_texto(valor)

            valores.append(valor)

        valores.append(id)

        set_sql = ", ".join([f"{coluna} = %s" for coluna in COLUNAS_EXPEDICAO])

        cur.execute(f"""
            UPDATE programacao_expedicao
            SET {set_sql}
            WHERE id = %s;
        """, valores)

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for("admin_programacao_expedicao"))

    cur.execute("""
        SELECT *
        FROM programacao_expedicao
        WHERE id = %s;
    """, (id,))

    registro = cur.fetchone()

    cur.close()
    conn.close()

    if not registro:
        return redirect(url_for("admin_programacao_expedicao"))

    return render_template(
        "editar_programacao_expedicao.html",
        registro=registro,
        colunas=COLUNAS_EXPEDICAO,
        colunas_numero=COLUNAS_NUMERO_EXPEDICAO,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/admin/programacao-cd/excluir/<int:id>", methods=["POST"])
def excluir_programacao_cd(id):
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM programacao_cd
        WHERE id = %s;
    """, (id,))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("admin_programacao_cd"))


@app.route("/admin/programacao-cd/limpar", methods=["POST"])
def limpar_programacao_cd():
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("DELETE FROM programacao_cd;")

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("admin_programacao_cd"))


@app.route("/admin/programacao-cross/excluir/<int:id>", methods=["POST"])
def excluir_programacao_cross(id):
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM programacao_cross
        WHERE id = %s;
    """, (id,))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("admin_programacao_cross"))


@app.route("/admin/programacao-cross/limpar", methods=["POST"])
def limpar_programacao_cross():
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("DELETE FROM programacao_cross;")

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("admin_programacao_cross"))


@app.route("/admin/programacao-expedicao/excluir/<int:id>", methods=["POST"])
def excluir_programacao_expedicao(id):
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM programacao_expedicao
        WHERE id = %s;
    """, (id,))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("admin_programacao_expedicao"))


@app.route("/admin/programacao-expedicao/limpar", methods=["POST"])
def limpar_programacao_expedicao():
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("DELETE FROM programacao_expedicao;")

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("admin_programacao_expedicao"))


@app.route("/admin/caminhoes/limpar", methods=["POST"])
def limpar_caminhoes():
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("DELETE FROM caminhoes;")

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("admin"))


@app.route("/cd/programacao")
def cd_programacao():
    if session.get("tipo") != "encarregado" and session.get("tipo") != "admin":
        return redirect(url_for("login"))

    if not is_cd_autorizado() and not is_admin():
        return redirect(url_for("encarregado"))

    busca = request.args.get("busca", "")
    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")

    programacoes = buscar_programacoes("programacao_cd", busca, data_inicio, data_fim)

    return render_template(
        "cd_programacao.html",
        programacoes=programacoes,
        busca=busca,
        data_inicio=data_inicio,
        data_fim=data_fim,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/cd/programacao-expedicao")
def cd_programacao_expedicao():
    if session.get("tipo") != "encarregado" and session.get("tipo") != "admin":
        return redirect(url_for("login"))

    if not is_cd_autorizado() and not is_admin():
        return redirect(url_for("encarregado"))

    busca = request.args.get("busca", "")
    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")

    programacoes = buscar_programacao_expedicao(busca, data_inicio, data_fim)

    return render_template(
        "cd_programacao_expedicao.html",
        programacoes=programacoes,
        busca=busca,
        data_inicio=data_inicio,
        data_fim=data_fim,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/cd/programacao-expedicao/editar-motorista/<int:id>", methods=["GET", "POST"])
def editar_motorista_expedicao(id):
    if not is_cd_autorizado():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        motorista = limpar_texto(request.form.get("motorista"))

        cur.execute("""
            UPDATE programacao_expedicao
            SET motorista = %s
            WHERE id = %s;
        """, (
            motorista,
            id
        ))

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for("cd_programacao_expedicao"))

    cur.execute("""
        SELECT *
        FROM programacao_expedicao
        WHERE id = %s;
    """, (id,))

    registro = cur.fetchone()

    cur.close()
    conn.close()

    if not registro:
        return redirect(url_for("cd_programacao_expedicao"))

    return render_template(
        "editar_motorista_expedicao.html",
        registro=registro,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/cross/programacao")
def cross_programacao():
    if session.get("tipo") != "encarregado" and session.get("tipo") != "admin":
        return redirect(url_for("login"))

    if not is_cross_autorizado() and not is_admin():
        return redirect(url_for("encarregado"))

    busca = request.args.get("busca", "")
    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")

    programacoes = buscar_programacoes("programacao_cross", busca, data_inicio, data_fim)

    return render_template(
        "cross_programacao.html",
        programacoes=programacoes,
        busca=busca,
        data_inicio=data_inicio,
        data_fim=data_fim,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/cross/registrar_chegada/<int:id>", methods=["POST"])
def registrar_chegada_cross(id):
    if not is_cross_autorizado():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        UPDATE programacao_cross
        SET horario_chegada = %s,
            usuario_chegada = %s
        WHERE id = %s;
    """, (
        agora_brasil(),
        session.get("usuario"),
        id
    ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("cross_programacao"))


@app.route("/cross/registrar_saida/<int:id>", methods=["POST"])
def registrar_saida_cross(id):
    if not is_cross_autorizado():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        UPDATE programacao_cross
        SET horario_saida = %s,
            usuario_saida = %s
        WHERE id = %s
          AND horario_chegada IS NOT NULL;
    """, (
        agora_brasil(),
        session.get("usuario"),
        id
    ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("cross_programacao"))


@app.route("/cross/registrar_justificativa/<int:id>", methods=["POST"])
def registrar_justificativa_cross(id):
    if not is_cross_autorizado():
        return redirect(url_for("login"))

    justificativa = limpar_texto(request.form.get("justificativa"))

    if not justificativa:
        return redirect(url_for("cross_programacao"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        UPDATE programacao_cross
        SET justificativa = %s,
            usuario_justificativa = %s,
            horario_justificativa = %s
        WHERE id = %s;
    """, (
        justificativa,
        session.get("usuario"),
        agora_brasil(),
        id
    ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("cross_programacao"))


@app.route("/exportar_programacao_cd")
def exportar_programacao_cd():
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM programacao_cd
        ORDER BY data_agenda_entrega ASC NULLS LAST, id DESC;
    """)

    dados = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Programacao CD"

    cabecalhos = ["ID"] + COLUNAS_PROGRAMACAO_CD + ["criado_em"]
    ws.append(cabecalhos)

    for item in dados:
        linha = [item.get("id")]

        for coluna in COLUNAS_PROGRAMACAO_CD:
            linha.append(item.get(coluna))

        linha.append(item.get("criado_em"))
        ws.append(linha)

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="programacao_cd.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/exportar_programacao_cross")
def exportar_programacao_cross():
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM programacao_cross
        ORDER BY data_agenda_entrega ASC NULLS LAST, id DESC;
    """)

    dados = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Programacao CROSS"

    cabecalhos = ["ID"] + COLUNAS_PROGRAMACAO + [
        "horario_chegada",
        "usuario_chegada",
        "horario_saida",
        "usuario_saida",
        "justificativa",
        "usuario_justificativa",
        "horario_justificativa",
        "criado_em"
    ]
    ws.append(cabecalhos)

    for item in dados:
        linha = [item.get("id")]

        for coluna in COLUNAS_PROGRAMACAO:
            linha.append(item.get(coluna))

        linha.append(item.get("horario_chegada"))
        linha.append(item.get("usuario_chegada"))
        linha.append(item.get("horario_saida"))
        linha.append(item.get("usuario_saida"))
        linha.append(item.get("justificativa"))
        linha.append(item.get("usuario_justificativa"))
        linha.append(item.get("horario_justificativa"))
        linha.append(item.get("criado_em"))

        ws.append(linha)

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="programacao_cross.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/exportar_programacao_expedicao")
def exportar_programacao_expedicao():
    if not is_admin():
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM programacao_expedicao
        ORDER BY id DESC;
    """)

    dados = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Programacao Expedicao"

    cabecalhos = ["ID"] + COLUNAS_EXPEDICAO + ["criado_em"]
    ws.append(cabecalhos)

    for item in dados:
        linha = [item.get("id")]

        for coluna in COLUNAS_EXPEDICAO:
            linha.append(item.get(coluna))

        linha.append(item.get("criado_em"))
        ws.append(linha)

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="programacao_expedicao.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/registros_encarregado")
def registros_encarregado():
    if session.get("tipo") != "admin":
        return redirect(url_for("login"))

    busca = request.args.get("busca", "")
    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")

    conn = conectar()
    cur = conn.cursor()

    query = """
        SELECT *
        FROM caminhoes
        WHERE horario_operacional IS NOT NULL
    """

    params = []

    if busca:
        query += """
            AND (
                placa ILIKE %s OR
                motorista ILIKE %s OR
                cpf ILIKE %s OR
                empresa ILIKE %s OR
                tipo_material ILIKE %s OR
                nota_fiscal ILIKE %s OR
                setor_doca ILIKE %s OR
                doca ILIKE %s OR
                equipe ILIKE %s
            )
        """
        termo = f"%{busca}%"
        params.extend([termo, termo, termo, termo, termo, termo, termo, termo, termo])

    if data_inicio:
        query += " AND horario_operacional::date >= %s"
        params.append(data_inicio)

    if data_fim:
        query += " AND horario_operacional::date <= %s"
        params.append(data_fim)

    query += " ORDER BY horario_operacional DESC;"

    cur.execute(query, params)
    registros = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "registros_encarregado.html",
        registros=registros,
        busca=busca,
        data_inicio=data_inicio,
        data_fim=data_fim,
        usuario=session.get("usuario"),
        perfil=session.get("tipo")
    )


@app.route("/relatorio_excel")
def relatorio_excel():
    if session.get("tipo") != "admin":
        return redirect(url_for("login"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM caminhoes
        ORDER BY id DESC;
    """)

    dados = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio PELOG"

    ws.append([
        "ID", "Placa", "Motorista", "CPF", "Empresa", "Material", "Nota Fiscal/DTS",
        "Setor", "Doca", "Status", "Entrada Portaria", "Autorizado Por",
        "Horário Autorização", "Início Doca", "Fim Doca",
        "Saida Portaria", "Saida Registrada Por",
        "Romaneio/Manifesto", "Produtos SKU", "Notas",
        "Quantidade NFS", "Paletes NF", "Paletes Conferido",
        "Peso KG", "Diferença Produtos", "Equipe", "Operacional Por",
        "Horário Operacional"
    ])

    for c in dados:
        ws.append([
            c.get("id"),
            c.get("placa"),
            c.get("motorista"),
            c.get("cpf"),
            c.get("empresa"),
            c.get("tipo_material"),
            c.get("nota_fiscal"),
            c.get("setor_doca"),
            c.get("doca"),
            c.get("status"),
            c.get("horario"),
            c.get("autorizado_por"),
            c.get("horario_autorizacao"),
            c.get("inicio_doca"),
            c.get("fim_doca"),
            c.get("horario_saida"),
            c.get("saida_registrada_por"),
            c.get("dts_observacao"),
            c.get("produtos_sku"),
            c.get("notas"),
            c.get("quantidade_nfs"),
            c.get("qtd_paletes_nf"),
            c.get("qtd_paletes_conferido"),
            c.get("peso_kg"),
            c.get("diferenca_produtos"),
            c.get("equipe"),
            c.get("operacional_cadastrado_por"),
            c.get("horario_operacional"),
        ])

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio_pelog.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/tv")
def tv():
    return render_template("tv.html")


@app.route("/dados_tv")
def dados_tv():
    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            id,
            placa,
            motorista,
            empresa,
            tipo_material,
            nota_fiscal,
            setor_doca,
            doca,
            status,
            inicio_doca
        FROM caminhoes
        WHERE status = 'na_doca'
          AND inicio_doca IS NOT NULL
        ORDER BY doca ASC, inicio_doca ASC;
    """)

    caminhoes = cur.fetchall()

    cur.close()
    conn.close()

    dados = []
    agora = agora_brasil()

    for c in caminhoes:
        tempo = 0

        if c.get("inicio_doca"):
            tempo = int((agora - c.get("inicio_doca")).total_seconds() / 60)

        dados.append({
            "id": c.get("id"),
            "placa": c.get("placa"),
            "motorista": c.get("motorista"),
            "empresa": c.get("empresa"),
            "tipo_material": c.get("tipo_material"),
            "nota_fiscal": c.get("nota_fiscal"),
            "setor_doca": c.get("setor_doca"),
            "doca": c.get("doca") or "SEM DOCA",
            "status": c.get("status"),
            "tempo": tempo
        })

    return jsonify(dados)


@app.route("/api/tv")
def api_tv():
    return dados_tv()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port, debug=True)
